"""
基础工具集 - 不依赖 LangChain

每个 Tool 只需提供：
- name: 工具名
- description: 工具描述（供 LLM 理解）
- invoke(args) -> str: 同步执行入口

为了与原 castorice_tools 保持兼容，这里同时提供与 LangChain BaseTool 一致的接口。
"""

import inspect
import os
import subprocess
import re
from typing import Dict, Any, List, Optional, Union, get_type_hints, get_origin, get_args

import httpx

from castorice.http_client import get_http_client
from castorice.logger import get_logger

_logger = get_logger(__name__)


def _get_httpx_client():
    """获取单例 httpx.Client（带浏览器 User-Agent，避免被 API 拦截）"""
    return get_http_client()


_SENSITIVE_FILE_PATTERNS = [
    ".env", ".env.", "id_rsa", "id_dsa", "id_ed25519",
    ".pem", ".ppk", "privkey",
]

_MAX_COMMAND_LENGTH = 4096
_MAX_CODE_LENGTH = 16384
_MAX_WRITE_CONTENT_LENGTH = 1048576  # 1 MB
_MAX_READ_LINES = 10000


def _default_allowed_paths() -> List[str]:
    """默认允许的路径列表：工作目录 + castorice_data + journals + ~/.castorice"""
    import os as _os
    paths = []
    cwd = _os.getcwd()
    paths.append(cwd)
    castorice_data = _os.path.join(cwd, "castorice_data")
    paths.append(castorice_data)
    journals_dir = _os.path.join(cwd, "journals")
    paths.append(journals_dir)
    dot_castorice = _os.path.join(_os.path.expanduser("~"), ".castorice")
    paths.append(dot_castorice)
    return paths


def _redirect_note_path(file_path: str) -> str:
    """将简单文件名的笔记(.txt/.md)重定向到 journals/ 目录。
    仅当路径是纯文件名（不含目录分隔符）且为笔记类扩展名时才重定向。
    """
    if not file_path:
        return file_path
    # 已经是绝对路径或包含子目录 → 不重定向
    if os.path.isabs(file_path) or "/" in file_path or "\\" in file_path:
        return file_path
    lower = file_path.lower()
    if lower.endswith(".txt") or lower.endswith(".md"):
        return os.path.join("journals", file_path)
    return file_path


def _is_path_safe(file_path: str, allowed_paths: Optional[List[str]]) -> bool:
    """
    检查文件路径是否安全（白名单模式）
    - 阻止路径遍历攻击（../ 或 ..\）
    - 阻止读取敏感文件
    - 默认使用 _default_allowed_paths() 作为白名单
    - 检查绝对路径是否在 allowed_paths 的任一目录下
    """
    if not file_path or not isinstance(file_path, str):
        return False

    raw_path = file_path.strip()

    if ".." in raw_path:
        return False

    abs_path = os.path.abspath(raw_path)
    canonical_path = os.path.realpath(raw_path)

    if abs_path != canonical_path:
        return False

    file_name = os.path.basename(abs_path).lower()

    for pattern in _SENSITIVE_FILE_PATTERNS:
        if pattern in file_name:
            return False

    if ".ssh" in abs_path.lower().replace("\\", "/").split("/"):
        return False

    # 禁止读取审计日志目录（防止 Agent 篡改/读取自身审计记录）
    path_parts = abs_path.lower().replace("\\", "/").split("/")
    if "audit_logs" in path_parts:
        return False

    if allowed_paths is None or len(allowed_paths) == 0:
        allowed_paths = _default_allowed_paths()

    for allowed in allowed_paths:
        abs_allowed = os.path.abspath(allowed)
        if abs_path == abs_allowed:
            return True
        if abs_path.startswith(abs_allowed.rstrip(os.sep) + os.sep):
            return True

    return False


class Tool:
    """极简工具基类（自研版，替代 LangChain BaseTool）"""

    # Python 类型 → JSON Schema 类型映射
    _TYPE_MAP = {
        str: "string",
        int: "integer",
        float: "number",
        bool: "boolean",
        list: "array",
        dict: "object",
    }

    def __init__(self, name: str, description: str, func, risk_level: str = "low"):
        self.name = name
        self.description = description
        self.func = func
        self.risk_level = risk_level

    def invoke(self, args: Dict[str, Any]) -> str:
        """执行工具，自动适配位置参数与关键字参数"""
        if isinstance(args, dict):
            return self.func(**args)
        return self.func(args)

    @classmethod
    def _resolve_type(cls, param_type) -> str:
        """
        P1-17: 解析 Python 类型为 JSON Schema 类型字符串，支持复合类型。

        - 基础类型 (str/int/float/bool) 直接查表
        - Optional[X] / Union[X, None] 解包取第一个非 None 类型
        - List[X] / Set / Tuple → "array"
        - Dict[K, V] → "object"
        - 无法识别 → "string"（兜底）
        """
        if param_type is None:
            return "string"
        # 基础类型直接查表
        if param_type in cls._TYPE_MAP:
            return cls._TYPE_MAP[param_type]
        # 复合类型
        origin = get_origin(param_type)
        if origin is not None:
            args = get_args(param_type)
            # Optional[X] = Union[X, None]，取第一个非 None 类型
            if origin is Union:
                non_none_args = [a for a in args if a is not type(None)]
                if non_none_args:
                    return cls._resolve_type(non_none_args[0])
                return "string"
            # List[X] / Set / Tuple → array
            if origin in (list, set, tuple, frozenset):
                return "array"
            # Dict[K, V] → object
            if origin is dict:
                return "object"
        return "string"

    @staticmethod
    def _extract_param_docs(func) -> Dict[str, str]:
        """
        P2-2: 从函数 docstring 提取参数描述。

        支持格式：
        - Sphinx: :param name: description
        - Google: Args: / Parameters: 块中的 "name: description"
        """
        doc = getattr(func, "__doc__", None)
        if not doc:
            return {}
        result: Dict[str, str] = {}
        # Sphinx 风格: :param name: description
        for m in re.finditer(r':param\s+(\w+)\s*:\s*(.+)', doc):
            result[m.group(1)] = m.group(2).strip()
        if result:
            return result
        # Google 风格: Args:/Parameters: 块
        google_block = re.search(
            r'(?:Args|Parameters)\s*:\s*\n((?:\s+\w+\s*:.*\n?)+)',
            doc,
        )
        if google_block:
            for line in google_block.group(1).strip().splitlines():
                m = re.match(r'\s+(\w+)\s*:\s*(.+)', line)
                if m:
                    result[m.group(1)] = m.group(2).strip()
        return result

    def to_openai_schema(self) -> dict:
        """生成 OpenAI Function Calling 格式的 tool schema"""
        try:
            hints = get_type_hints(self.func)
            sig = inspect.signature(self.func)
        except (TypeError, ValueError):
            return self._minimal_schema()

        properties = {}
        required = []
        # P2-2: 从 docstring 提取参数描述
        param_docs = self._extract_param_docs(self.func)

        for param_name, param in sig.parameters.items():
            if param_name in ("self", "cls", "args", "kwargs"):
                continue

            param_type = hints.get(param_name)
            # P1-17: 使用 _resolve_type 处理复合类型
            json_type = self._resolve_type(param_type)

            # P2-2: 优先用 docstring 描述，其次用参数名+类型标注
            desc = param_docs.get(param_name, "")
            if not desc:
                desc = param_name
                if param_type is int:
                    desc = f"{param_name} (整数)"
                elif param_type is float:
                    desc = f"{param_name} (数字)"
                elif param_type is bool:
                    desc = f"{param_name} (布尔值)"

            prop = {"type": json_type, "description": desc}
            properties[param_name] = prop

            # 没有默认值的参数是必填的
            if param.default is inspect.Parameter.empty:
                required.append(param_name)

        if not properties:
            return self._minimal_schema()

        return {
            "type": "function",
            "function": {
                "name": self.name,
                "description": self.description,
                "parameters": {
                    "type": "object",
                    "properties": properties,
                    "required": required,
                },
            },
        }

    def to_anthropic_schema(self) -> dict:
        """生成 Anthropic tool_use 格式的 tool schema"""
        openai_schema = self.to_openai_schema()
        func = openai_schema["function"]
        return {
            "name": func["name"],
            "description": func["description"],
            "input_schema": func["parameters"],
        }

    def _minimal_schema(self) -> dict:
        """最小 schema（无法推导参数类型时的兜底）"""
        return {
            "type": "function",
            "function": {
                "name": self.name,
                "description": self.description,
                "parameters": {
                    "type": "object",
                    "properties": {},
                },
            },
        }


_registered_tools: Dict[str, Tool] = {}


def register_tool(name: str, description: str, risk_level: str = "low"):
    """
    工具注册装饰器
    - 装饰函数，自动创建 Tool 实例并存入 _registered_tools
    - 保留原有函数不变
    - risk_level: 审计风险等级 (low / medium / high)
    """
    def decorator(func):
        tool = Tool(name=name, description=description, func=func, risk_level=risk_level)
        _registered_tools[name] = tool
        return func
    return decorator


def reload_tools():
    """重新加载工具注册表（热更新）。

    清空当前注册表并重新导入 base_tools / web_tools 模块代码，
    触发 @register_tool 装饰器重新注册工具。
    保持同一 _registered_tools dict 对象引用，确保其他模块持有的引用仍然有效。
    """
    import importlib
    _registered_tools.clear()
    # 重新导入 base_tools 以触发装饰器重新注册
    import castorice.tools.base_tools as _bt
    importlib.reload(_bt)
    # reload 会创建新的 _registered_tools，需同步回当前 dict 以保持引用有效
    _registered_tools.update(_bt._registered_tools)
    # 重新导入 web_tools（外部信息检索工具）
    try:
        import castorice.tools.web_tools as _wt
        importlib.reload(_wt)
        _registered_tools.update(_wt._registered_tools)
    except (ImportError, ModuleNotFoundError, AttributeError) as e:
        _logger.warning(f"重新加载 web_tools 失败: {e}")
    # 重新导入 eigenflux_tool（EigenFlux 网络工具）
    try:
        import castorice.tools.eigenflux_tool as _et
        importlib.reload(_et)
        _registered_tools.update(_et._registered_tools)
    except (ImportError, ModuleNotFoundError, AttributeError) as e:
        _logger.warning(f"重新加载 eigenflux_tool 失败: {e}")


# ========== 1. 联网搜索 ==========
@register_tool(
    name="web_search",
    description=(
        "联网搜索信息（基于 DuckDuckGo，无需 API Key）。\n"
        "参数:\n"
        "- query (必填, str): 搜索关键词，建议 2-5 个词\n"
        "- max_results (可选, int, 默认 5): 返回结果数量，建议 3-10\n"
        "适用场景：查新闻、查资料、查人物、查事件、查最新动态\n"
        "不适用：查实时天气（用 get_weather）、查股价（用 stock_price）、查视频（用 youtube/bilibili 搜索）\n"
        "Few-shot 示例：\n"
        "  例1: query='2026年人工智能发展趋势' → 返回行业趋势文章\n"
        "  例2: query='小米SU7发布会' max_results=3 → 限制返回3条\n"
        "  例3: query='Python asyncio 教程' → 返回教程链接"
    ),
)
def _web_search(query: str, max_results: int = 5) -> str:
    """使用 DuckDuckGo 搜索并返回结果摘要"""
    try:
        from ddgs import DDGS
    except ImportError:
        try:
            from duckduckgo_search import DDGS
        except ImportError:
            return "未安装搜索库，请执行: pip install ddgs"

    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(query, max_results=max_results))
        if not results:
            return "未搜索到结果"
        lines = []
        for i, r in enumerate(results, 1):
            title = r.get("title", "")
            snippet = r.get("body", "") or r.get("snippet", "")
            href = r.get("href", "") or r.get("url", "")
            lines.append(f"{i}. {title}\n   {snippet}\n   {href}")
        return "\n".join(lines)
    except (httpx.HTTPError, ConnectionError, TimeoutError) as e:
        return f"搜索失败: {e}"


# ========== 1.5 实时天气查询 ==========

_WEATHER_DESC_ZH = {
    "Sunny": "晴",
    "Clear": "晴",
    "Partly cloudy": "多云",
    "Partly Cloudy": "多云",
    "Cloudy": "阴",
    "Overcast": "阴天",
    "Mist": "薄雾",
    "Fog": "雾",
    "Freezing fog": "冻雾",
    "Patchy rain possible": "可能有阵雨",
    "Patchy rain nearby": "局部有小雨",
    "Patchy snow possible": "可能有阵雪",
    "Patchy sleet possible": "可能有雨夹雪",
    "Patchy freezing drizzle possible": "可能有冻毛毛雨",
    "Thundery outbreaks possible": "可能有雷暴",
    "Blowing snow": "吹雪",
    "Blizzard": "暴风雪",
    "Light drizzle": "小毛毛雨",
    "Patchy light drizzle": "零星小毛毛雨",
    "Freezing drizzle": "冻毛毛雨",
    "Heavy freezing drizzle": "强冻毛毛雨",
    "Light rain": "小雨",
    "Moderate rain at times": "间歇性中雨",
    "Moderate rain": "中雨",
    "Heavy rain at times": "间歇性大雨",
    "Heavy rain": "大雨",
    "Light freezing rain": "小冻雨",
    "Moderate or heavy freezing rain": "中到大雨冻雨",
    "Light sleet": "小雨夹雪",
    "Moderate or heavy sleet": "中到大雨夹雪",
    "Light snow": "小雪",
    "Patchy light snow": "零星小雪",
    "Moderate snow": "中雪",
    "Patchy moderate snow": "间歇性中雪",
    "Heavy snow": "大雪",
    "Patchy heavy snow": "间歇性大雪",
    "Ice pellets": "冰粒",
    "Light rain shower": "小阵雨",
    "Moderate or heavy rain shower": "中到大雨阵雨",
    "Torrential rain shower": "暴雨",
    "Light sleet showers": "小雨夹雪阵雨",
    "Moderate or heavy sleet showers": "中到大雨夹雪阵雨",
    "Light snow showers": "小阵雪",
    "Moderate or heavy snow showers": "中到大雪阵雪",
    "Light showers of ice pellets": "小冰粒阵雨",
    "Moderate or heavy showers of ice pellets": "中到大冰粒阵雨",
    "Patchy light rain with thunder": "零星小雨伴雷暴",
    "Moderate or heavy rain with thunder": "中到大雨伴雷暴",
    "Patchy light snow with thunder": "零星小雪伴雷暴",
    "Moderate or heavy snow with thunder": "中到大雪伴雷暴",
    "Smoky haze": "雾霾",
    "Smoke": "烟霾",
    "Haze": "霾",
}


def _weather_zh(text: str) -> str:
    """天气描述英译中（不区分大小写）"""
    if not text:
        return text
    lower = text.strip().lower()
    for en, zh in _WEATHER_DESC_ZH.items():
        if en.lower() == lower:
            return zh
    return text


@register_tool(
    name="get_weather",
    description=(
        "查询城市实时天气和未来 7 天预报（基于 wttr.in 免费 API，无需 API Key）。\n"
        "参数:\n"
        "- city (必填, str): 城市名，支持中英文，如 '大连'、'北京'、'Shanghai'\n"
        "- day (可选, int, 默认 0): 预报天数，0=今天实时天气，1=明天，2=后天...最多 7 天\n"
        "返回内容包含：实时温度、体感温度、天气状况（中文）、湿度、风速、未来 7 天预报\n"
        "Few-shot 示例：\n"
        "  例1: city='大连' → 大连今天实时天气 + 未来 7 天\n"
        "  例2: city='北京' day=1 → 北京明天天气预报\n"
        "  例3: city='上海' day=3 → 上海 3 天后天气预报"
    ),
)
def _get_weather(city: str, day: int = 0, lang: str = "zh") -> str:
    """
    查询实时天气和未来7天预报（基于 wttr.in 免费 API，无需 API Key）
    比 web_search 更准确、更快，返回实时气温、天气状况、风速等
    """
    import urllib.parse
    from datetime import datetime, timedelta

    try:
        encoded_city = urllib.parse.quote(city)
        url = f"https://wttr.in/{encoded_city}?format=j1&lang={lang}"

        client = _get_httpx_client()
        resp = client.get(url, headers={"User-Agent": "castorice-agent/2.0"})
        resp.raise_for_status()
        data = resp.json()

        # 限制 day 参数范围
        day = max(0, min(day, 7))

        # 获取指定日期的预报
        if day < len(data["weather"]):
            target_day = data["weather"][day]
        else:
            target_day = data["weather"][-1]

        # 获取今天的实时天气
        current = data["current_condition"][0]
        weather_today = data["weather"][0]

        temp_c = current.get("temp_C", "N/A")
        feels_like = current.get("FeelsLikeC", "N/A")

        desc = ""
        lang_key = f"lang_{lang}"
        en_desc = current.get("weatherDesc", [{}])[0].get("value", "")
        if lang_key in current and current[lang_key]:
            lang_desc = current[lang_key][0].get("value", "")
            if lang_desc and lang_desc.lower() != en_desc.lower():
                desc = lang_desc
        if not desc:
            desc = _weather_zh(en_desc)

        humidity = current.get("humidity", "N/A")
        wind_speed = current.get("windspeedKmph", "N/A")
        wind_dir = current.get("winddir16Point", "")

        maxtemp = weather_today.get("maxtempC", "N/A")
        mintemp = weather_today.get("mintempC", "N/A")

        # 获取指定日期的详细信息
        date_str = target_day.get("date", "")
        target_max = target_day.get("maxtempC", "?")
        target_min = target_day.get("mintempC", "?")
        hourly = target_day.get("hourly", [])
        midday = hourly[len(hourly) // 2] if hourly else {}

        target_desc = ""
        target_en_desc = midday.get("weatherDesc", [{}])[0].get("value", "")
        if lang_key in midday and midday[lang_key]:
            target_lang_desc = midday[lang_key][0].get("value", "")
            if target_lang_desc and target_lang_desc.lower() != target_en_desc.lower():
                target_desc = target_lang_desc
        if not target_desc:
            target_desc = _weather_zh(target_en_desc)

        # 生成日期描述
        day_labels = ["今天", "明天", "后天", "大后天", "4天后", "5天后", "6天后", "7天后"]
        day_label = day_labels[day] if day < len(day_labels) else f"{day}天后"

        # 如果查询的是今天，显示实时天气
        if day == 0:
            result = (
                f"【{city} 实时天气】\n"
                f"当前温度: {temp_c}°C（体感 {feels_like}°C）\n"
                f"天气状况: {desc}\n"
                f"今日气温: {mintemp}°C ~ {maxtemp}°C\n"
                f"湿度: {humidity}%\n"
                f"风速: {wind_speed} km/h {wind_dir}"
            )
        else:
            result = (
                f"【{city} {day_label} ({date_str}) 天气预报】\n"
                f"天气状况: {target_desc}\n"
                f"气温: {target_min}°C ~ {target_max}°C"
            )

        # 添加未来7天预报（无论查询哪一天，都显示完整预报）
        forecast_lines = []
        for i, day_data in enumerate(data["weather"][:7]):
            date = day_data.get("date", "")
            max_c = day_data.get("maxtempC", "?")
            min_c = day_data.get("mintempC", "?")
            hourly_data = day_data.get("hourly", [])
            midday_data = hourly_data[len(hourly_data) // 2] if hourly_data else {}

            day_desc = ""
            day_en_desc = midday_data.get("weatherDesc", [{}])[0].get("value", "")
            if lang_key in midday_data and midday_data[lang_key]:
                day_lang_desc = midday_data[lang_key][0].get("value", "")
                if day_lang_desc and day_lang_desc.lower() != day_en_desc.lower():
                    day_desc = day_lang_desc
            if not day_desc:
                day_desc = _weather_zh(day_en_desc)

            label = day_labels[i] if i < len(day_labels) else f"{i}天后"
            forecast_lines.append(f"{label} ({date}): {day_desc}, {min_c}°C ~ {max_c}°C")

        result += (
            f"\n\n【未来7天预报】\n"
            + "\n".join(forecast_lines)
        )

        return result
    except (httpx.HTTPError, ConnectionError, TimeoutError) as e:
        return f"天气查询失败: {e}"


# ========== 2. 读文件 ==========
@register_tool(
    name="read_file",
    description=(
        "读取文本文件内容（自动 UTF-8 解码，失败时回退到 ignore）。\n"
        "参数:\n"
        "- file_path (必填, str): 文件绝对或相对路径\n"
        "- max_lines (可选, int, 默认 200): 最多读取行数（超出截断）\n"
        "- allowed_paths (可选, list[str]): 允许访问的目录白名单（受配置约束）\n"
        "安全限制：自动阻止读取 .env、.ssh/id_rsa、.pem、id_dsa 等敏感文件\n"
        "Few-shot 示例：\n"
        "  例1: file_path='README.md' → 读取项目说明\n"
        "  例2: file_path='src/main.py' max_lines=50 → 只读前 50 行\n"
        "  例3: file_path='C:/Users/sheng/notes.txt' → 读取绝对路径"
    ),
    risk_level="medium",  # P2-3: 读取文件内容属于中等风险
)

def _read_file(file_path: str, max_lines: int = 200, allowed_paths: Optional[List[str]] = None) -> str:
    try:
        file_path = _redirect_note_path(file_path)
        if max_lines > _MAX_READ_LINES:
            return f"[BLOCKED] max_lines 超过上限 {_MAX_READ_LINES}"
        if not _is_path_safe(file_path, allowed_paths):
            return f"[BLOCKED] 文件路径不在白名单中或为敏感文件: {file_path}"

        path = os.path.abspath(file_path)
        if not os.path.exists(path):
            return f"文件不存在: {path}"
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
        lines = content.splitlines()
        if len(lines) > max_lines:
            content = "\n".join(lines[:max_lines]) + f"\n... (截断,共 {len(lines)} 行)"
        return content
    except (OSError, IOError, PermissionError) as e:
        return f"读取失败: {e}"


# ========== 3. 写文件 ==========
@register_tool(
    name="write_file",
    description=(
        "将文本内容写入文件（覆盖或新建，UTF-8 编码）。\n"
        "参数:\n"
        "- file_path (必填, str): 目标文件路径\n"
        "- content (必填, str): 要写入的文本内容\n"
        "- allowed_paths (可选, list[str]): 允许写入的目录白名单\n"
        "安全限制：自动阻止写入 .env、.ssh 等敏感路径；写入前会做内容审查（file_guard）\n"
        "Few-shot 示例：\n"
        "  例1: file_path='note.txt' content='Hello' → 写入文件\n"
        "  例2: file_path='logs/output.log' content='日志内容' → 追加到日志\n"
        "  例3: file_path='config.json' content='{\"key\": \"value\"}' → 写 JSON 配置"
    ),
    risk_level="high",
)
def _write_file(file_path: str, content: str, allowed_paths: Optional[List[str]] = None) -> str:
    try:
        file_path = _redirect_note_path(file_path)
        if len(content) > _MAX_WRITE_CONTENT_LENGTH:
            return f"[BLOCKED] 内容长度超过上限 {_MAX_WRITE_CONTENT_LENGTH} 字节"
        if not _is_path_safe(file_path, allowed_paths):
            return f"[BLOCKED] 文件路径不在白名单中或为敏感文件: {file_path}"

        from castorice.security.file_guard import get_file_guard
        guard = get_file_guard()
        allowed, reason = guard.check_write_allowed(file_path, content)
        if not allowed:
            return f"[BLOCKED] {reason}"

        os.makedirs(os.path.dirname(os.path.abspath(file_path)) or ".", exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)
        return f"已写入 {len(content)} 字符到 {file_path}"
    except (OSError, IOError, PermissionError) as e:
        return f"写入失败: {e}"


# ========== 4. 终端命令 ==========

_TERMINAL_WHITELIST = {
    "ls", "dir", "cd", "pwd", "echo", "cat", "type", "grep", "findstr",
    "wc", "head", "tail", "python", "python3", "pip", "pip3", "git",
    "npm", "node", "curl", "wget", "whoami", "date", "time", "hostname",
    "ipconfig", "ifconfig", "ping", "tree", "more", "less", "sort",
    "uniq", "awk", "sed", "cut", "paste", "join", "split", "tr",
    "base64", "md5sum", "sha256sum", "du", "df", "free", "top", "ps",
    "tasklist",
}

_QUOTED_STRIP_RE = re.compile(r'"[^"]*"|\'[^\']*\'')


def _detect_command_injection(cmd: str) -> bool:
    """
    智能命令注入检测（分层检测，避免误拦合法命令）。

    返回 True 表示检测到注入，False 表示安全。

    检测层次（按优先级）：
    L0: 换行符/回车符 — shell 解释为命令分隔符，无条件拦截
    L1: 管道/重定向（|, >, <）— 无条件拦截（排除引号内）
    L2: 命令替换 $(...) 和 ${...} — 全字符串检测（双引号内也生效）
    L3: 反引号命令替换 `command` — 仅当反引号间有实际内容时拦截
    L4: 命令链接（;, &, &&）— 仅当两侧都有命令词时拦截

    合法放行示例：
    - "echo $PATH" — 简单变量引用
    - "echo hello; world" — 单侧分号，无链接命令
    - "python -c \"print(1)\"" — 括号在引号内
    """
    stripped = cmd.strip()
    if not stripped:
        return False

    # L0: 换行符/回车符 — shell 解释为命令分隔符，无条件拦截
    if '\n' in stripped or '\r' in stripped:
        return True

    # L0: 剥离引号内容，避免误拦引号内的特殊字符
    # 例如 echo "hello > world" → 剥离后只剩 echo
    unquoted = _QUOTED_STRIP_RE.sub('', stripped)

    # L1: 管道和重定向 — 无条件拦截（在非引号部分出现即为高风险）
    if re.search(r'[|><]', unquoted):
        return True

    # L2: 命令替换模式 — 全字符串检测（双引号内也能执行替换）
    # $(...) — 经典命令替换（如 echo "$(whoami)"）
    if '$(' in stripped:
        return True

    # ${...} — 变量扩展可触发命令执行（如 echo "${cmd}"）
    if '${' in stripped:
        return True

    # L3: 反引号命令替换 `command` — 仅当反引号对中有实际内容时拦截
    # 单反引号（如 Markdown 代码片段）放行；`...` 有内容则拦截
    if '`' in stripped:
        if re.search(r'`[^`]+`', stripped):
            return True

    # L4: 命令链接运算符 — 仅当两侧都有命令词时拦截
    # ; — 仅拦截 "cmd1; cmd2" 形式的链接；放行尾部/头部/独立的 ;
    if re.search(r'\S+\s*;\s*\S+', unquoted):
        return True

    # & 或 && — 仅拦截 "cmd1 & cmd2" 形式的链接
    if re.search(r'\S+\s*&&?\s*\S+', unquoted):
        return True

    return False


@register_tool(
    name="terminal",
    description=(
        "执行 shell 命令（Windows/PowerShell/cmd，白名单安全限制）。\n"
        "参数:\n"
        "- command (必填, str): shell 命令字符串\n"
        "- timeout (可选, int, 默认 30): 超时时间（秒）\n"
        "安全限制：\n"
        "  - 只能执行白名单内命令（ls/dir/cd/echo/python/git/curl/wget 等 49 个）\n"
        "  - 阻止危险命令（format/del/rm/shutdown/mkfs/dd/chown 等）\n"
        "  - 智能拦截命令注入（管道/重定向/命令替换/链接）\n"
        "Few-shot 示例：\n"
        "  例1: command='ls' → 列出当前目录\n"
        "  例2: command='pip list' → 查看已安装包\n"
        "  例3: command='git status' → 查看 git 状态"
    ),
    risk_level="high",
)
def _terminal(command: str, timeout: int = 30) -> str:
    """执行 shell 命令（白名单安全限制 + 智能命令注入防护）"""
    stripped = command.strip()
    if not stripped:
        return "[BLOCKED] 命令不能为空"
    if len(stripped) > _MAX_COMMAND_LENGTH:
        return f"[BLOCKED] 命令长度超过上限 {_MAX_COMMAND_LENGTH} 字节"

    from castorice.security.file_guard import get_file_guard
    guard = get_file_guard()
    allowed, reason = guard.check_command_allowed(command)
    if not allowed:
        return f"[BLOCKED] {reason}"

    cmd_parts = stripped.split()
    cmd_prefix = cmd_parts[0].lower() if cmd_parts else ""

    if cmd_prefix not in _TERMINAL_WHITELIST:
        _logger.warning(f"TERMINAL BLOCKED: 命令不在白名单中: {cmd_prefix}")
        return f"[BLOCKED] 命令不在白名单中: {cmd_prefix}"

    # 拦截可执行子代码命令和脚本文件执行，防止绕过沙盒执行任意代码
    # 注意：pip/pip3 不在此列，因为 pip 的 -e 是 editable 安装参数（pip install -e .），非代码执行
    _CODE_EXEC_PREFIXES = {"python", "python3", "node", "ruby", "perl"}
    _CODE_EXEC_FLAGS = {"-c", "-e", "-m"}
    _SCRIPT_EXTENSIONS = {".py", ".js", ".rb", ".pl", ".sh", ".bash"}
    if cmd_prefix in _CODE_EXEC_PREFIXES:
        # 检查是否有代码执行参数（-c/-e/-m）
        for part in cmd_parts[1:]:
            flag = part.split("=", 1)[0].lower()
            if flag in _CODE_EXEC_FLAGS:
                _logger.warning(
                    f"TERMINAL BLOCKED: 检测到子代码执行参数 {flag}，"
                    f"请使用 python_repl 工具: {stripped}"
                )
                return (
                    f"[BLOCKED] 检测到子代码执行参数 {flag}，"
                    f"禁止通过 {cmd_prefix} 直接执行代码，请使用 python_repl 工具"
                )
        # 检查是否直接执行脚本文件（防止 python script.py 绕过 AST 沙箱）
        for part in cmd_parts[1:]:
            if part.startswith("-"):
                continue
            lower_part = part.lower()
            if any(lower_part.endswith(ext) for ext in _SCRIPT_EXTENSIONS):
                _logger.warning(
                    f"TERMINAL BLOCKED: 检测到脚本文件执行 {part}，"
                    f"请使用 python_repl 工具: {stripped}"
                )
                return (
                    f"[BLOCKED] 检测到脚本文件执行 {part}，"
                    f"禁止通过 {cmd_prefix} 直接执行脚本，请使用 python_repl 工具"
                )

    # SSRF 防护：拦截 curl/wget 访问内部/私有网络地址
    _NETWORK_CMDS = {"curl", "wget"}
    if cmd_prefix in _NETWORK_CMDS:
        try:
            from .web_tools import _is_internal_url
            url_candidates = []
            for part in cmd_parts[1:]:
                if part.startswith("-"):
                    continue
                if part.lower().startswith(("http://", "https://")):
                    url_candidates.append(part)
                elif "://" in part:
                    url_candidates.append(part)
            for url in url_candidates:
                if _is_internal_url(url):
                    _logger.warning(f"TERMINAL BLOCKED: {cmd_prefix} 访问内部地址被拦截: {url}")
                    return f"[BLOCKED] SSRF 防护：禁止通过 {cmd_prefix} 访问内部/私有网络地址: {url}"
        except (ImportError, ModuleNotFoundError, AttributeError) as _e:
            _logger.warning(f"TERMINAL SSRF 检查异常，默认拦截: {_e}")
            return f"[BLOCKED] SSRF 检查异常，已默认拦截: {_e}"

    if _detect_command_injection(stripped):
        _logger.warning(f"TERMINAL BLOCKED: 检测到命令注入: {stripped}")
        return "[BLOCKED] 检测到命令注入"

    # P1: 黑名单命令检查 —— 防止白名单内的命令配合换行符绕过后执行危险命令
    cmd_lower = stripped.lower()
    for bl_cmd in _DANGEROUS_COMMANDS_BLACKLIST:
        if bl_cmd in cmd_parts or bl_cmd in cmd_lower.split():
            _logger.warning(f"TERMINAL BLOCKED: 危险命令: {bl_cmd}")
            return f"[BLOCKED] 危险命令被拦截: {bl_cmd}"

    # P1: pip install 安全检查 —— 防止安装恶意包
    if cmd_prefix in ("pip", "pip3") and len(cmd_parts) >= 2 and cmd_parts[1] == "install":
        _logger.warning(f"TERMINAL BLOCKED: pip install 需要确认: {stripped}")
        return (
            "[BLOCKED] pip install 操作需要人工确认。"
            "安装第三方包可能引入安全风险（恶意 setup.py），"
            "请手动执行 pip install 命令。"
        )

    # P2: git clone SSRF 防护
    if cmd_prefix == "git" and len(cmd_parts) >= 2 and cmd_parts[1] == "clone":
        try:
            from .web_tools import _is_internal_url
            for part in cmd_parts[2:]:
                if part.startswith("-"):
                    continue
                if part.lower().startswith(("http://", "https://", "git://")) or "://" in part:
                    if _is_internal_url(part):
                        _logger.warning(f"TERMINAL BLOCKED: git clone 访问内部地址: {part}")
                        return f"[BLOCKED] SSRF 防护：禁止 git clone 内部/私有网络地址: {part}"
                    break
        except (ImportError, ModuleNotFoundError, AttributeError):
            pass

    try:
        result = subprocess.run(
            command, shell=True, capture_output=True, text=True,
            timeout=timeout, encoding="utf-8", errors="ignore",
        )
        out = result.stdout or ""
        err = result.stderr or ""
        if result.returncode != 0:
            return f"exit={result.returncode}\nstdout:\n{out}\nstderr:\n{err}"
        return out if out else "(无输出)"
    except subprocess.TimeoutExpired:
        return f"[TIMEOUT] 命令执行超过 {timeout} 秒"
    except (OSError, ValueError, RuntimeError) as e:
        return f"执行失败: {e}"


# ========== 4.5 命令安全检查（公开 API） ==========
# 注意：黑名单只保留命令首词，因为 cmd_prefix 仅取 cmd_parts[0]。
# 多词命令（如 "taskkill /f"、"net stop"）永远不会被首词匹配，因此拆分为单词。
_DANGEROUS_COMMANDS_BLACKLIST = {
    # Windows 危险命令
    "format", "del", "rd", "rmdir", "rm", "reg", "regedit",
    "diskpart", "bcdedit", "cipher", "sfc",
    # 系统控制
    "shutdown", "restart", "logoff", "tsshutdown", "psshutdown",
    "taskkill", "net",
    # 文件破坏
    "move", "ren", "rename", "copy", "xcopy", "robocopy",
    "attrib", "takeown", "icacls",
    # 网络危险
    "netsh", "firewall", "route", "arp",
    # Linux 危险命令
    "mkfs", "dd", "fdisk", "parted", "mount", "umount",
    "useradd", "userdel", "passwd", "chown", "chmod",
    "iptables", "ip6tables", "tc",
    # 其他 —— shell 解释器首词（含 -c / -e 等子代码执行参数）
    "powershell", "cmd", "bash", "sh", "zsh", "ksh",
}


def is_command_safe(command: str) -> bool:
    """
    检查命令是否安全（公开 API，供测试和其他模块使用）

    参数：
        command: 待检查的命令

    返回：True 表示安全，False 表示危险
    """
    if not command or not isinstance(command, str):
        return False

    stripped = command.strip()
    if not stripped:
        return False

    # 提取命令前缀
    cmd_parts = stripped.split()
    if not cmd_parts:
        return False

    cmd_prefix = cmd_parts[0].lower()
    # 去掉路径前缀（如 C:\Windows\System32\format.exe -> format）
    if "\\" in cmd_prefix or "/" in cmd_prefix:
        cmd_prefix = cmd_prefix.replace("\\", "/").split("/")[-1]
    # 去掉 .exe 后缀
    if cmd_prefix.endswith(".exe"):
        cmd_prefix = cmd_prefix[:-4]

    # 检查黑名单
    if cmd_prefix in _DANGEROUS_COMMANDS_BLACKLIST:
        return False

    # 检查命令注入
    if _detect_command_injection(stripped):
        return False

    return True


# ========== 5. Python REPL ==========

_SAFE_BUILTINS = {
    "print": print,
    "len": len,
    "range": range,
    "str": str,
    "int": int,
    "float": float,
    "list": list,
    "dict": dict,
    "sum": sum,
    "min": min,
    "max": max,
    "abs": abs,
    "round": round,
    "sorted": sorted,
    "enumerate": enumerate,
    "zip": zip,
    "map": map,
    "filter": filter,
    "isinstance": isinstance,
    "bool": bool,
    "set": set,
    "tuple": tuple,
    "reversed": reversed,
    "all": all,
    "any": any,
    "ord": ord,
    "chr": chr,
    "hex": hex,
    "bin": bin,
    "format": format,
    "pow": pow,
    "divmod": divmod,
    "next": next,
    "iter": iter,
    "slice": slice,
    "True": True,
    "False": False,
    "None": None,
    "Exception": Exception,
    "ValueError": ValueError,
    "TypeError": TypeError,
    "IndexError": IndexError,
    "KeyError": KeyError,
    "StopIteration": StopIteration,
    "NotImplementedError": NotImplementedError,
    "RuntimeError": RuntimeError,
    "AttributeError": AttributeError,
}


_DANGEROUS_PATTERNS = [
    "__import__", "__subclasses__", "__bases__", "__globals__", "__code__",
    "__class__", "__base__", "__mro__", "__dict__", "__getattribute__",
    "open(", "exec(", "eval(", "compile(", "subprocess", "os.", "sys.",
    "socket.", "urllib.", "http.", "requests.", "shutil.", "pathlib.",
    "getattr(", "globals(", "locals(", "vars(",
    "itertools.count", "itertools.cycle",
]


def _is_code_safe_ast(code: str) -> tuple:
    """
    AST 级安全扫描：检测危险名称、属性、import 语句、getattr 调用、无限循环。
    可防御字符串拼接绕过（如 "ope" + "n("）。
    """
    import ast
    try:
        tree = ast.parse(code)
    except SyntaxError as e:
        return False, f"语法错误: {e}"

    dangerous_names = {
        "__import__", "__subclasses__", "__bases__", "__globals__",
        "__code__", "__class__", "__base__", "__mro__", "__dict__",
        "__getattribute__",
        "open", "exec", "eval", "compile", "subprocess",
        "os", "sys", "socket", "urllib", "http", "requests", "shutil",
        "pathlib", "importlib", "builtins", "__builtins__",
        "getattr", "globals", "locals", "vars", "type",
    }

    for node in ast.walk(tree):
        if isinstance(node, (ast.Import, ast.ImportFrom)):
            return False, "禁止 import 语句"
        if isinstance(node, ast.Name):
            if node.id in dangerous_names:
                return False, f"检测到危险名称: {node.id}"
        if isinstance(node, ast.Attribute):
            if node.attr in dangerous_names:
                return False, f"检测到危险属性: {node.attr}"
        # 检测 getattr() 调用 —— 可用于沙盒逃逸
        if isinstance(node, ast.Call):
            func = node.func
            if isinstance(func, ast.Name) and func.id == "getattr":
                return False, "检测到 getattr 调用，禁止使用"
            if isinstance(func, ast.Name) and func.id == "type" and len(node.args) >= 3:
                return False, "检测到 type() 三参数形式（动态创建类），禁止使用"
        # 检测 while True / while 1 无限循环
        if isinstance(node, ast.While):
            test = node.test
            if isinstance(test, ast.Constant) and test.value is True:
                return False, "检测到 while True 无限循环"
            if isinstance(test, ast.Constant) and isinstance(test.value, int) and test.value != 0:
                return False, "检测到 while <非零常量> 无限循环"
            if isinstance(test, ast.Name) and test.id == "True":
                return False, "检测到 while True 无限循环"

    return True, ""


@register_tool(
    name="python_repl",
    description=(
        "执行 Python 代码片段（受限沙箱环境，48 个安全内置函数）。\n"
        "参数:\n"
        "- code (必填, str): Python 代码字符串\n"
        "- timeout (可选, int, 默认 30): 超时时间（秒）\n"
        "安全限制：\n"
        "  - 禁止 import 语句、__import__、open、exec、eval、compile、subprocess 等\n"
        "  - 禁止访问 os、sys、socket、shutil、pathlib 等系统模块\n"
        "  - 双层防护：字符串模式匹配 + AST 语法树扫描\n"
        "可用内置：print/len/range/str/int/list/dict/sum/min/max/sorted/enumerate/zip 等\n"
        "Few-shot 示例：\n"
        "  例1: code='print(sum(range(10)))' → 输出 45\n"
        "  例2: code='x = [1,2,3]; print(sum(x), len(x))' → 输出 6 3\n"
        "  例3: code='result = sorted([3,1,2]); print(result)' → 输出 [1, 2, 3]"
    ),
    risk_level="medium",
)
def _python_repl(code: str, timeout: int = 30) -> str:
    """
    安全受限的 Python 代码执行沙箱

    特性：
    - 使用白名单内置函数，无文件系统和网络访问
    - 无 __import__、open、exec、eval 等危险函数
    - 保留 stdout 重定向捕获输出功能
    - 双层防护：字符串模式匹配 + AST 语法树扫描
    - 软超时：通过 threading.Timer 检测超时（无法中断已运行的 exec）
    """
    if len(code) > _MAX_CODE_LENGTH:
        return f"[BLOCKED] 代码长度超过上限 {_MAX_CODE_LENGTH} 字节"
    code_lower = code.lower()
    for pattern in _DANGEROUS_PATTERNS:
        if pattern in code_lower:
            return f"[安全拦截] 检测到危险代码模式: {pattern}"

    safe_ast, ast_reason = _is_code_safe_ast(code)
    if not safe_ast:
        return f"[安全拦截] {ast_reason}"

    # P0: 集成自我保护系统 — 检测自我毁灭代码
    try:
        from castorice.security.self_protection import SelfProtectionSystem
        sp = SelfProtectionSystem()
        if sp.detect_self_destruction(code):
            return "[安全拦截] 自我保护系统检测到自我毁灭倾向，已拦截"
    except (OSError, ImportError, RuntimeError) as e:
        _logger.warning(f"自我保护系统不可用: {e}")

    try:
        import sys
        import threading
        from io import StringIO

        old_stdout = sys.stdout
        sys.stdout = StringIO()

        # 软超时检测：threading.Timer 无法真正中断 exec，但可记录超时状态
        # 配合 AST 层对 while True 的拦截，可防止绝大多数无限循环
        timed_out = False

        def _timeout_handler():
            nonlocal timed_out
            timed_out = True

        timer = threading.Timer(timeout, _timeout_handler)
        timer.start()
        try:
            safe_globals = {"__builtins__": _SAFE_BUILTINS}
            exec(code, safe_globals, {})
        finally:
            timer.cancel()
            output = sys.stdout.getvalue()
            sys.stdout = old_stdout

        if timed_out:
            return f"Error: 代码执行超过 {timeout} 秒（软超时）"
        return output if output else "(无输出)"
    except (OSError, ValueError, RuntimeError) as e:
        # 恢复 stdout（异常路径）
        try:
            sys.stdout = old_stdout
        except Exception:
            pass
        return f"执行出错: {type(e).__name__}: {e}"


# ========== 6. 文档读取（PDF/DOCX/XLSX） ==========
@register_tool(
    name="read_document",
    description=(
        "读取 PDF/Word/Excel 文档内容（自动识别扩展名）。\n"
        "参数:\n"
        "- file_path (必填, str): 文档绝对路径，支持 .pdf/.docx/.xlsx\n"
        "- allowed_paths (可选, list[str]): 允许访问的目录白名单\n"
        "支持格式：\n"
        "  - .pdf: 使用 pypdf 提取文本（最多 5000 字符）\n"
        "  - .docx: 使用 python-docx 读取段落（最多 5000 字符）\n"
        "  - .xlsx: 使用 openpyxl 读取工作表（最多 5000 字符）\n"
        "  - 其他: 尝试用 read_file 读取（按文本文件处理）\n"
        "Few-shot 示例：\n"
        "  例1: file_path='report.pdf' → 提取 PDF 文本\n"
        "  例2: file_path='contract.docx' → 读取 Word 文档\n"
        "  例3: file_path='sales_data.xlsx' → 读取 Excel 表格"
    ),
    risk_level="medium",
)
def _read_document(file_path: str, allowed_paths: Optional[List[str]] = None) -> str:
    """读取 PDF/Word/Excel 文档"""
    if not _is_path_safe(file_path, allowed_paths):
        return f"[BLOCKED] 文件路径不在白名单中或为敏感文件: {file_path}"
    if not os.path.exists(file_path):
        return f"文件不存在: {file_path}"
    ext = os.path.splitext(file_path)[1].lower()
    try:
        if ext == ".pdf":
            from pypdf import PdfReader
            reader = PdfReader(file_path)
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
            return text[:5000]
        if ext == ".docx":
            from docx import Document
            doc = Document(file_path)
            return "\n".join(p.text for p in doc.paragraphs)[:5000]
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join(str(c) if c is not None else "" for c in row))
            return "\n".join(lines)[:5000]
        return _read_file(file_path, allowed_paths=allowed_paths)
    except (OSError, IOError, PermissionError, ImportError, ModuleNotFoundError, ValueError, RuntimeError) as e:
        return f"文档读取失败: {e}"


# ========== 7. 获取当前时间 ==========
@register_tool(
    name="get_current_time",
    description=(
        "获取当前日期和时间（包含时区信息）。\n"
        "无需参数。\n"
        "返回内容：UTC 时间、本地时间、星期、年份、月份、日期\n"
        "Few-shot 示例：\n"
        "  例1: 调用 → '当前时间（UTC）: 2026-07-22 09:30:00 UTC\\n当前时间（本地）: ...\\n星期: 周三'"
    ),
)
def _get_current_time() -> str:
    """获取当前日期时间，支持多种格式输出"""
    from datetime import datetime, timezone
    
    now = datetime.now(timezone.utc)
    local_now = datetime.now().astimezone()
    
    return (
        f"当前时间（UTC）: {now.strftime('%Y-%m-%d %H:%M:%S')} UTC\n"
        f"当前时间（本地）: {local_now.strftime('%Y-%m-%d %H:%M:%S')}\n"
        f"星期: {['周一', '周二', '周三', '周四', '周五', '周六', '周日'][local_now.weekday()]}\n"
        f"年份: {local_now.year}\n"
        f"月份: {local_now.month}月\n"
        f"日期: {local_now.day}日"
    )


# ========== 工具注册入口 ==========
def get_base_tools(config: Optional[Dict[str, Any]] = None) -> List[Tool]:
    """
    获取所有基础工具实例列表
    - 优先从 _registered_tools 获取已注册的工具
    - 根据配置决定启用哪些工具
    - 从配置中读取 allowed_paths 传给文件读写工具
    - 保持向后兼容
    - 自动加载 web_tools 中的外部信息检索工具
    """
    # 自动导入 web_tools 以注册外部信息检索工具
    try:
        from . import web_tools  # noqa: F401
    except ImportError:
        pass

    # 自动导入 eigenflux_tool 以注册 EigenFlux 网络工具
    try:
        from . import eigenflux_tool  # noqa: F401
    except ImportError:
        pass

    allowed_paths = None
    if config:
        tools_cfg = config.get("tools", {})
        # 从 read_file 和 write_file 配置中读取 allowed_paths
        read_cfg = tools_cfg.get("read_file", {})
        write_cfg = tools_cfg.get("write_file", {})
        if isinstance(read_cfg, dict):
            allowed_paths = read_cfg.get("allowed_paths", None)
        if allowed_paths is None and isinstance(write_cfg, dict):
            allowed_paths = write_cfg.get("allowed_paths", None)

    # 读取各工具的 enabled 配置
    tools_enabled = {}
    if config:
        tools_cfg = config.get("tools", {})
        for key in ["web_search", "get_weather", "get_current_time", "read_file", "write_file", "terminal", "python_repl", "read_document",
                    "web_fetch", "wikipedia_search", "arxiv_search", "news_search",
                    "github_search", "youtube_search", "bilibili_search",
                    "ip_info", "stock_price", "translate_text",
                    "anime_search", "anime_season",
                    "vrchat_search", "vrchat_popular_worlds",
                    "vrchat_user_status", "vrchat_world_info",
                    "generate_image", "analyze_image", "extract_text_from_image",
                    "pixiv_search", "pixiv_popular", "pixiv_user_works",
                    "ef_feed"]:
            tool_cfg = tools_cfg.get(key, {})
            if isinstance(tool_cfg, dict):
                tools_enabled[key] = tool_cfg.get("enabled", True)
            else:
                tools_enabled[key] = True

    def is_enabled(name: str) -> bool:
        return tools_enabled.get(name, True)

    all_tools = []

    if is_enabled("web_search") and "web_search" in _registered_tools:
        all_tools.append(_registered_tools["web_search"])

    if is_enabled("get_weather") and "get_weather" in _registered_tools:
        all_tools.append(_registered_tools["get_weather"])

    if is_enabled("get_current_time") and "get_current_time" in _registered_tools:
        all_tools.append(_registered_tools["get_current_time"])

    if is_enabled("read_file") and "read_file" in _registered_tools:
        base_tool = _registered_tools["read_file"]
        if allowed_paths is not None:
            def _read_file_wrapper(file_path: str, max_lines: int = 200) -> str:
                return base_tool.func(file_path, max_lines=max_lines, allowed_paths=allowed_paths)
            all_tools.append(Tool(
                name="read_file",
                description=base_tool.description,
                func=_read_file_wrapper,
            ))
        else:
            all_tools.append(base_tool)

    if is_enabled("write_file") and "write_file" in _registered_tools:
        base_tool = _registered_tools["write_file"]
        if allowed_paths is not None:
            def _write_file_wrapper(file_path: str, content: str) -> str:
                return base_tool.func(file_path, content, allowed_paths=allowed_paths)
            all_tools.append(Tool(
                name="write_file",
                description=base_tool.description,
                func=_write_file_wrapper,
            ))
        else:
            all_tools.append(base_tool)

    if is_enabled("terminal") and "terminal" in _registered_tools:
        all_tools.append(_registered_tools["terminal"])

    if is_enabled("python_repl") and "python_repl" in _registered_tools:
        all_tools.append(_registered_tools["python_repl"])

    if is_enabled("read_document") and "read_document" in _registered_tools:
        base_tool = _registered_tools["read_document"]
        if allowed_paths is not None:
            def _read_document_wrapper(file_path: str) -> str:
                return base_tool.func(file_path, allowed_paths=allowed_paths)
            all_tools.append(Tool(
                name="read_document",
                description=base_tool.description,
                func=_read_document_wrapper,
                risk_level=base_tool.risk_level,
            ))
        else:
            all_tools.append(base_tool)

    # 外部信息检索工具
    for name in ["web_fetch", "wikipedia_search", "arxiv_search", "news_search",
                 "github_search", "youtube_search", "bilibili_search",
                 "ip_info", "stock_price", "translate_text",
                 "anime_search", "anime_season",
                 "vrchat_search", "vrchat_popular_worlds",
                 "vrchat_user_status", "vrchat_world_info",
                 "generate_image", "analyze_image", "extract_text_from_image",
                 "pixiv_search", "pixiv_popular", "pixiv_user_works"]:
        if is_enabled(name) and name in _registered_tools:
            all_tools.append(_registered_tools[name])

    # 所有 EigenFlux 工具（动态收集，ef_ 开头的全部启用）
    for name in sorted(_registered_tools.keys()):
        if name.startswith("ef_") and is_enabled(name):
            all_tools.append(_registered_tools[name])

    return all_tools
